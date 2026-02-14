# Poll Kiosk Backend - Flask Application
# Supports surveys (multi-question), active survey queue, session tracking

from flask import Flask, render_template, request, jsonify, make_response, send_file
from functools import wraps
import json
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database import Database

app = Flask(__name__)

# ---------------------------------------------------------------- config

db_path = os.path.join('data', 'polls.db') if os.path.exists('data') else 'polls.db'
db = Database(db_path)

def load_config():
    if os.path.exists('config.json'):
        with open('config.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        'admin_username': 'admin',
        'admin_password': 'changeme'
    }

def save_config(config):
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

config = load_config()

# ---------------------------------------------------------- auth helpers

def check_auth(username, password):
    return username == config['admin_username'] and password == config['admin_password']

def authenticate():
    return make_response(
        'Необходима авторизация', 401,
        {'WWW-Authenticate': 'Basic realm="Admin Area"'}
    )

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

# ------------------------------------------------------------- public pages

@app.route('/')
def index():
    return render_template('index.html')

# ---------------------------------------------------------------- public API
#
# Session model (client-side):
#   The frontend maintains a "session" of survey IDs it is currently working
#   through.  On first load (or after reset) it fetches /api/session-config
#   which returns the current ordered list of active survey IDs + the full
#   content of all their polls.  The frontend iterates questions locally and
#   only contacts the server for:
#     - POST /api/vote            (record an answer)
#     - GET  /api/session-config  (detect changes between sessions)
#
#   Change detection: the frontend polls /api/session-config every 5 s.
#   The response includes a "version" hash (list of active survey IDs joined).
#   When the version changes the frontend:
#     - If the user is mid-session (answered ≥1 question): finishes current
#       session, then switches.
#     - Otherwise: switches immediately.

@app.route('/api/session-config', methods=['GET'])
def session_config():
    """
    Returns the current active survey queue.
    Response:
    {
        "version": "<str>",            # changes when active surveys change
        "surveys": [
            {
                "id": <int>,
                "title": "<str>",
                "polls": [
                    {"id": <int>, "question": "<str>", "answers": [...]}
                ]
            }
        ]
    }
    """
    surveys = db.get_active_surveys()
    version = ','.join(str(s['id']) for s in surveys)
    # strip created_at / position from polls to keep payload lean
    clean = []
    for s in surveys:
        clean.append({
            'id': s['id'],
            'title': s['title'],
            'show_title': s.get('show_title', True),
            'polls': [
                {
                    'id': p['id'],
                    'question': p['question'],
                    'answers': p['answers'],
                    'multi_select': p.get('multi_select', False)
                }
                for p in s['polls']
            ]
        })
    return jsonify({'version': version, 'surveys': clean})


@app.route('/api/vote', methods=['POST'])
def submit_vote():
    """
    Submit a vote.
    Single choice:  { "poll_id": <int>, "answer_index": <int>, "session_id": <str> }
    Multi-select:   { "poll_id": <int>, "answer_indices": [<int>, ...], "session_id": <str> }
    """
    data = request.get_json()
    if not data or 'poll_id' not in data:
        return jsonify({'error': 'Invalid request'}), 400

    poll_id = data['poll_id']
    session_id = data.get('session_id') or None

    poll = db.get_poll(poll_id)
    if not poll:
        return jsonify({'error': 'Poll not found'}), 404

    # Collect indices from either answer_index or answer_indices
    if 'answer_indices' in data:
        indices = data['answer_indices']
        if not isinstance(indices, list) or len(indices) == 0:
            return jsonify({'error': 'answer_indices must be a non-empty list'}), 400
    elif 'answer_index' in data:
        indices = [data['answer_index']]
    else:
        return jsonify({'error': 'answer_index or answer_indices required'}), 400

    for idx in indices:
        if not isinstance(idx, int) or idx < 0 or idx >= len(poll['answers']):
            return jsonify({'error': f'Invalid answer index: {idx}'}), 400

    db.save_vote(poll_id, indices, request.remote_addr, session_id=session_id)
    return jsonify({'success': True})


# ---------------------------------------------------------------- admin pages

@app.route('/admin')
@requires_auth
def admin_panel():
    return render_template('admin.html')


# ---------------------------------------------------------------- admin API — polls (questions)

@app.route('/api/admin/polls', methods=['GET'])
@requires_auth
def get_polls():
    polls = db.get_all_polls()
    return jsonify({'polls': polls})


@app.route('/api/admin/polls', methods=['POST'])
@requires_auth
def create_poll():
    data = request.get_json()
    if not data or 'question' not in data or 'answers' not in data:
        return jsonify({'error': 'Invalid request'}), 400

    question = data['question'].strip()
    answers = [a.strip() for a in data['answers'] if a.strip()]
    if not question or len(answers) < 2:
        return jsonify({'error': 'Question and at least 2 answers required'}), 400

    multi_select = bool(data.get('multi_select', False))
    poll_id = db.create_poll(question, answers, multi_select=multi_select)
    return jsonify({'success': True, 'poll_id': poll_id})


@app.route('/api/admin/polls/<int:poll_id>', methods=['PUT'])
@requires_auth
def update_poll(poll_id):
    data = request.get_json()
    if not data or 'question' not in data or 'answers' not in data:
        return jsonify({'error': 'Invalid request'}), 400
    question = data['question'].strip()
    answers = [a.strip() for a in data['answers'] if a.strip()]
    if not question or len(answers) < 2:
        return jsonify({'error': 'Question and at least 2 answers required'}), 400
    multi_select = data.get('multi_select')  # None = don't change
    if multi_select is not None:
        multi_select = bool(multi_select)
    db.update_poll(poll_id, question, answers, multi_select=multi_select)
    return jsonify({'success': True})


@app.route('/api/admin/polls/<int:poll_id>', methods=['DELETE'])
@requires_auth
def delete_poll(poll_id):
    db.delete_poll(poll_id)
    return jsonify({'success': True})


@app.route('/api/admin/stats/<int:poll_id>', methods=['GET'])
@requires_auth
def get_poll_stats(poll_id):
    poll = db.get_poll(poll_id)
    if not poll:
        return jsonify({'error': 'Poll not found'}), 404
    stats = db.get_poll_stats(poll_id)
    return jsonify({'poll': poll, 'stats': stats})


# ---------------------------------------------------------------- admin API — surveys

@app.route('/api/admin/surveys', methods=['GET'])
@requires_auth
def get_surveys():
    surveys = db.get_all_surveys()
    active_ids = db.get_active_survey_ids()
    return jsonify({'surveys': surveys, 'active_survey_ids': active_ids})


@app.route('/api/admin/surveys', methods=['POST'])
@requires_auth
def create_survey():
    data = request.get_json()
    if not data or 'title' not in data or 'poll_ids' not in data:
        return jsonify({'error': 'Invalid request'}), 400

    title = data['title'].strip()
    poll_ids = data['poll_ids']
    show_title = data.get('show_title', True)
    if not title:
        return jsonify({'error': 'Title required'}), 400

    survey_id = db.create_survey(title, poll_ids, show_title=show_title)
    return jsonify({'success': True, 'survey_id': survey_id})


@app.route('/api/admin/surveys/<int:survey_id>', methods=['PUT'])
@requires_auth
def update_survey(survey_id):
    data = request.get_json()
    if not data or 'title' not in data or 'poll_ids' not in data:
        return jsonify({'error': 'Invalid request'}), 400

    title = data['title'].strip()
    poll_ids = data['poll_ids']
    show_title = data.get('show_title', None)
    if not title:
        return jsonify({'error': 'Title required'}), 400

    db.update_survey(survey_id, title, poll_ids, show_title=show_title)
    return jsonify({'success': True})


@app.route('/api/admin/surveys/<int:survey_id>', methods=['DELETE'])
@requires_auth
def delete_survey(survey_id):
    db.delete_survey(survey_id)
    return jsonify({'success': True})


@app.route('/api/admin/surveys/<int:survey_id>/votes', methods=['DELETE'])
@requires_auth
def reset_survey_votes(survey_id):
    """Delete all votes for all polls in a survey (for clearing test data)."""
    ok = db.reset_survey_votes(survey_id)
    if not ok:
        return jsonify({'error': 'Survey not found'}), 404
    return jsonify({'success': True})


@app.route('/api/admin/surveys/<int:survey_id>', methods=['GET'])
@requires_auth
def get_survey(survey_id):
    survey = db.get_survey(survey_id)
    if not survey:
        return jsonify({'error': 'Survey not found'}), 404
    return jsonify(survey)


@app.route('/api/admin/surveys/<int:survey_id>/stats', methods=['GET'])
@requires_auth
def get_survey_stats(survey_id):
    stats = db.get_survey_stats(survey_id)
    if stats is None:
        return jsonify({'error': 'Survey not found'}), 404
    return jsonify({'stats': stats})


# ---------------------------------------------------------------- Excel helpers

# Shared style constants
_BOLD       = Font(bold=True)
_WHITE_BOLD = Font(bold=True, color='FFFFFF')
_FILL_SURVEY = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_FILL_Q      = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
_FILL_META   = PatternFill(start_color='2D2150', end_color='2D2150', fill_type='solid')
_CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT_WRAP   = Alignment(horizontal='left', vertical='center', wrap_text=True)
META_COLS    = 2  # № and Дата


def _build_col_map(polls):
    """
    Build a column map: for each poll, list of (col_index, header, answer_index_or_None).
    Single-choice polls occupy 1 column (answer_index_or_None = None → put text).
    Multi-select polls occupy N columns, one per answer option (answer_index_or_None = i).
    Returns (col_map, total_data_cols) where col_map is
      { poll_id: [(ci_offset, header, answer_idx), ...] }
    and ci_offset is 0-based offset from the first data column.
    """
    col_map = {}
    offset = 0
    for poll in polls:
        pid = poll['id']
        if poll.get('multi_select'):
            cols = []
            for i, ans in enumerate(poll['answers']):
                cols.append((offset, ans, i))
                offset += 1
        else:
            cols = [(offset, poll['question'], None)]
            offset += 1
        col_map[pid] = cols
    return col_map, offset  # offset == total data columns


def _write_survey_block(ws, survey_title, polls, rows, start_row):
    """
    Write one survey block (title row + header rows + data rows) into worksheet
    starting at start_row.  Returns the next free row number after the block.

    Column layout for each question:
      Single-choice  → 1 column: text of chosen answer
      Multi-select   → N columns (one per answer option): '✓' or ''
    """
    col_map, total_data_cols = _build_col_map(polls)
    total_cols = META_COLS + total_data_cols
    r = start_row

    # ── Row 1: survey title ─────────────────────────────────────────────────
    c = ws.cell(row=r, column=1, value=f'Опрос: {survey_title}')
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = _FILL_SURVEY
    c.alignment = _CENTER
    if total_cols > 1:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    r += 1

    # ── Row 2: question-group header (merge cells for multi_select) ──────────
    # For single-choice: one cell with question text.
    # For multi-select: merge N cells, write question text as group header.
    for ci, label in enumerate(['№', 'Дата'], start=1):
        hc = ws.cell(row=r, column=ci, value=label)
        hc.font = _WHITE_BOLD
        hc.fill = _FILL_META
        hc.alignment = _CENTER

    for poll in polls:
        pid = poll['id']
        cols = col_map[pid]
        first_ci = META_COLS + cols[0][0] + 1
        last_ci  = META_COLS + cols[-1][0] + 1
        hc = ws.cell(row=r, column=first_ci, value=poll['question'])
        hc.font = _BOLD
        hc.fill = _FILL_Q
        hc.alignment = _CENTER if len(cols) > 1 else _LEFT_WRAP
        if len(cols) > 1:
            ws.merge_cells(start_row=r, start_column=first_ci, end_row=r, end_column=last_ci)

    ws.row_dimensions[r].height = 40
    r += 1

    # ── Row 3 (only for multi_select polls): sub-column answer-option headers ─
    has_multi = any(poll.get('multi_select') for poll in polls)
    if has_multi:
        # meta cells — empty but styled same as header
        for ci in range(1, META_COLS + 1):
            hc = ws.cell(row=r, column=ci, value='')
            hc.fill = _FILL_META

        for poll in polls:
            pid = poll['id']
            for (off, header, ans_idx) in col_map[pid]:
                ci = META_COLS + off + 1
                hc = ws.cell(row=r, column=ci, value=header)
                hc.font = _BOLD
                if ans_idx is None:
                    # single-choice sub-header: repeat question (already merged above)
                    hc.fill = _FILL_Q
                    hc.alignment = _LEFT_WRAP
                else:
                    # multi-select sub-option header
                    hc.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                    hc.alignment = _LEFT_WRAP
        ws.row_dimensions[r].height = 35
        r += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, resp in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=ri).alignment = _CENTER
        ws.cell(row=r, column=2, value=resp['voted_at']).alignment = _CENTER
        for poll in polls:
            pid = poll['id']
            ans = resp['answers'].get(pid)
            for (off, header, ans_idx) in col_map[pid]:
                ci = META_COLS + off + 1
                if ans is None:
                    ws.cell(row=r, column=ci, value='').alignment = _LEFT_WRAP
                elif ans_idx is None:
                    # single-choice: put text of chosen answer
                    text = ans['answer_texts'][0] if ans['answer_texts'] else ''
                    ws.cell(row=r, column=ci, value=text).alignment = _LEFT_WRAP
                else:
                    # multi-select: '✓' if this option was chosen
                    chosen = ans_idx in ans['answer_indices']
                    ws.cell(row=r, column=ci, value='✓' if chosen else '').alignment = _CENTER
        r += 1

    # ── Column widths (only set when wider than current) ────────────────────
    from openpyxl.utils import get_column_letter
    col_a = ws.column_dimensions['A']
    if col_a.width < 5:
        col_a.width = 5
    col_b = ws.column_dimensions['B']
    if col_b.width < 18:
        col_b.width = 18
    for poll in polls:
        pid = poll['id']
        for (off, header, ans_idx) in col_map[pid]:
            ci = META_COLS + off + 1
            letter = get_column_letter(ci)
            desired = min(len(header) + 4, 40)
            cur = ws.column_dimensions[letter].width or 0
            if desired > cur:
                ws.column_dimensions[letter].width = desired

    return r  # next free row


@app.route('/api/admin/surveys/<int:survey_id>/export', methods=['GET'])
@requires_auth
def export_survey_excel(survey_id):
    data = db.get_survey_respondents(survey_id)
    if data is None:
        return jsonify({'error': 'Survey not found'}), 404
    survey = db.get_survey(survey_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ответы'

    _write_survey_block(ws, survey['title'], data['polls'], data['rows'], start_row=1)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = survey['title'].replace('/', '-').replace('\\', '-')
    filename = f'survey_{survey_id}_{safe_title}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/admin/export', methods=['GET'])
@requires_auth
def export_all_surveys_excel():
    """Export all surveys into one Excel sheet, blocks separated by an empty row."""
    surveys = db.get_all_surveys()
    if not surveys:
        return jsonify({'error': 'No surveys found'}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Все опросы'

    current_row = 1
    for idx, s in enumerate(surveys):
        data = db.get_survey_respondents(s['id'])
        if data is None:
            continue
        current_row = _write_survey_block(ws, s['title'], data['polls'], data['rows'],
                                          start_row=current_row)
        current_row += 1  # empty separator row between surveys

    # Freeze only the very first header row of first block
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    filename = f'all_surveys_{date.today().isoformat()}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ---------------------------------------------------------------- admin API — active surveys queue

@app.route('/api/admin/active-surveys', methods=['GET'])
@requires_auth
def get_active_surveys():
    ids = db.get_active_survey_ids()
    return jsonify({'active_survey_ids': ids})


@app.route('/api/admin/active-surveys', methods=['POST'])
@requires_auth
def set_active_surveys():
    """
    Replace active survey queue.
    Body: { "survey_ids": [<int>, ...] }   # ordered list; empty list deactivates all
    """
    data = request.get_json()
    if data is None or 'survey_ids' not in data:
        return jsonify({'error': 'Invalid request'}), 400

    survey_ids = data['survey_ids']

    # Validate all IDs exist
    for sid in survey_ids:
        if not db.get_survey(sid):
            return jsonify({'error': f'Survey {sid} not found'}), 404

    db.set_active_surveys(survey_ids)
    return jsonify({'success': True})


# ---------------------------------------------------------------- entry point

if __name__ == '__main__':
    if not os.path.exists('config.json'):
        save_config(config)
        print("Created default config.json")
        print("Default admin credentials: admin / changeme")

    if not os.path.exists('data'):
        os.makedirs('data')
        print("Created data directory")

    app.run(host='0.0.0.0', port=5000, debug=False)
