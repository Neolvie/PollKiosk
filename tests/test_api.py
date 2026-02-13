"""
Automated tests for PollKiosk backend API.
Run with:  python -m pytest tests/test_api.py -v
"""

import json
import os
import sys
import tempfile

import pytest

# Ensure the app module can be imported from the repo root
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app import app as flask_app
from database import Database


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def db_file():
    """Create a temporary SQLite database and clean it up afterwards."""
    fd, path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    yield path
    os.unlink(path)


@pytest.fixture
def client(db_file, monkeypatch):
    """Flask test client wired to a fresh in-memory DB."""
    import app as app_module

    # Replace the global db instance with a fresh one
    test_db = Database(db_file)
    monkeypatch.setattr(app_module, 'db', test_db)

    flask_app.config['TESTING'] = True
    with flask_app.test_client() as c:
        yield c


def auth_headers():
    """HTTP Basic auth header for admin (default credentials)."""
    import base64
    creds = base64.b64encode(b'admin:changeme').decode()
    return {'Authorization': f'Basic {creds}'}


# ---------------------------------------------------------------------------
# Helper shortcuts
# ---------------------------------------------------------------------------

def create_poll(client, question='Q1', answers=('A', 'B')):
    r = client.post('/api/admin/polls',
                    json={'question': question, 'answers': list(answers)},
                    headers=auth_headers())
    assert r.status_code == 200
    return r.get_json()['poll_id']


def create_survey(client, title='Survey 1', poll_ids=None, show_title=True):
    r = client.post('/api/admin/surveys',
                    json={'title': title, 'poll_ids': poll_ids or [], 'show_title': show_title},
                    headers=auth_headers())
    assert r.status_code == 200
    return r.get_json()['survey_id']


# ===========================================================================
# POLLS
# ===========================================================================

class TestPolls:

    def test_create_poll(self, client):
        pid = create_poll(client, 'Favourite colour?', ['Red', 'Blue', 'Green'])
        assert isinstance(pid, int)

    def test_get_polls(self, client):
        create_poll(client)
        r = client.get('/api/admin/polls', headers=auth_headers())
        assert r.status_code == 200
        data = r.get_json()
        assert len(data['polls']) == 1
        assert data['polls'][0]['question'] == 'Q1'

    def test_create_poll_requires_two_answers(self, client):
        r = client.post('/api/admin/polls',
                        json={'question': 'Bad', 'answers': ['Only one']},
                        headers=auth_headers())
        assert r.status_code == 400

    def test_update_poll(self, client):
        pid = create_poll(client, 'Old question', ['A', 'B'])
        r = client.put(f'/api/admin/polls/{pid}',
                       json={'question': 'New question', 'answers': ['X', 'Y', 'Z']},
                       headers=auth_headers())
        assert r.status_code == 200

        # Verify via GET
        r2 = client.get('/api/admin/polls', headers=auth_headers())
        poll = r2.get_json()['polls'][0]
        assert poll['question'] == 'New question'
        assert poll['answers'] == ['X', 'Y', 'Z']

    def test_delete_poll(self, client):
        pid = create_poll(client)
        r = client.delete(f'/api/admin/polls/{pid}', headers=auth_headers())
        assert r.status_code == 200

        r2 = client.get('/api/admin/polls', headers=auth_headers())
        assert len(r2.get_json()['polls']) == 0

    def test_poll_requires_auth(self, client):
        r = client.get('/api/admin/polls')
        assert r.status_code == 401


# ===========================================================================
# SURVEYS
# ===========================================================================

class TestSurveys:

    def test_create_survey_empty(self, client):
        sid = create_survey(client, 'Empty survey')
        assert isinstance(sid, int)

    def test_create_survey_with_polls(self, client):
        pid1 = create_poll(client, 'Q1', ['A', 'B'])
        pid2 = create_poll(client, 'Q2', ['C', 'D'])
        sid = create_survey(client, 'Full survey', [pid1, pid2])

        r = client.get(f'/api/admin/surveys/{sid}', headers=auth_headers())
        data = r.get_json()
        assert data['title'] == 'Full survey'
        assert len(data['polls']) == 2
        assert data['polls'][0]['question'] == 'Q1'
        assert data['polls'][1]['question'] == 'Q2'

    def test_show_title_flag_default_true(self, client):
        sid = create_survey(client, 'Survey', show_title=True)
        r = client.get(f'/api/admin/surveys/{sid}', headers=auth_headers())
        assert r.get_json()['show_title'] is True

    def test_show_title_flag_false(self, client):
        sid = create_survey(client, 'Hidden title survey', show_title=False)
        r = client.get(f'/api/admin/surveys/{sid}', headers=auth_headers())
        assert r.get_json()['show_title'] is False

    def test_update_survey_title_and_show_title(self, client):
        pid = create_poll(client)
        sid = create_survey(client, 'Old name', [pid], show_title=True)

        r = client.put(f'/api/admin/surveys/{sid}',
                       json={'title': 'New name', 'poll_ids': [pid], 'show_title': False},
                       headers=auth_headers())
        assert r.status_code == 200

        r2 = client.get(f'/api/admin/surveys/{sid}', headers=auth_headers())
        data = r2.get_json()
        assert data['title'] == 'New name'
        assert data['show_title'] is False

    def test_delete_survey(self, client):
        sid = create_survey(client)
        r = client.delete(f'/api/admin/surveys/{sid}', headers=auth_headers())
        assert r.status_code == 200

        r2 = client.get(f'/api/admin/surveys/{sid}', headers=auth_headers())
        assert r2.status_code == 404

    def test_get_all_surveys_returns_show_title(self, client):
        create_survey(client, 'S1', show_title=True)
        create_survey(client, 'S2', show_title=False)
        r = client.get('/api/admin/surveys', headers=auth_headers())
        surveys = r.get_json()['surveys']
        assert len(surveys) == 2
        by_title = {s['title']: s for s in surveys}
        assert by_title['S1']['show_title'] is True
        assert by_title['S2']['show_title'] is False


# ===========================================================================
# ACTIVE QUEUE
# ===========================================================================

class TestActiveQueue:

    def test_set_and_get_active_surveys(self, client):
        sid1 = create_survey(client, 'First')
        sid2 = create_survey(client, 'Second')

        r = client.post('/api/admin/active-surveys',
                        json={'survey_ids': [sid1, sid2]},
                        headers=auth_headers())
        assert r.status_code == 200

        r2 = client.get('/api/admin/active-surveys', headers=auth_headers())
        assert r2.get_json()['active_survey_ids'] == [sid1, sid2]

    def test_clear_active_surveys(self, client):
        sid = create_survey(client)
        client.post('/api/admin/active-surveys',
                    json={'survey_ids': [sid]},
                    headers=auth_headers())
        client.post('/api/admin/active-surveys',
                    json={'survey_ids': []},
                    headers=auth_headers())

        r = client.get('/api/admin/active-surveys', headers=auth_headers())
        assert r.get_json()['active_survey_ids'] == []

    def test_set_active_unknown_survey_returns_404(self, client):
        r = client.post('/api/admin/active-surveys',
                        json={'survey_ids': [9999]},
                        headers=auth_headers())
        assert r.status_code == 404


# ===========================================================================
# SESSION CONFIG (public)
# ===========================================================================

class TestSessionConfig:

    def test_empty_config(self, client):
        r = client.get('/api/session-config')
        assert r.status_code == 200
        data = r.get_json()
        assert data['surveys'] == []
        assert data['version'] == ''

    def test_config_contains_show_title(self, client):
        pid = create_poll(client, 'Q?', ['Yes', 'No'])
        sid = create_survey(client, 'My Survey', [pid], show_title=False)
        client.post('/api/admin/active-surveys',
                    json={'survey_ids': [sid]},
                    headers=auth_headers())

        r = client.get('/api/session-config')
        data = r.get_json()
        assert len(data['surveys']) == 1
        survey = data['surveys'][0]
        assert survey['show_title'] is False
        assert survey['title'] == 'My Survey'
        assert len(survey['polls']) == 1
        assert survey['polls'][0]['question'] == 'Q?'

    def test_version_changes_when_queue_changes(self, client):
        r1 = client.get('/api/session-config')
        v1 = r1.get_json()['version']

        sid = create_survey(client)
        client.post('/api/admin/active-surveys',
                    json={'survey_ids': [sid]},
                    headers=auth_headers())

        r2 = client.get('/api/session-config')
        v2 = r2.get_json()['version']

        assert v1 != v2


# ===========================================================================
# VOTING
# ===========================================================================

class TestVoting:

    def test_submit_vote(self, client):
        pid = create_poll(client, 'Like it?', ['Yes', 'No'])
        r = client.post('/api/vote', json={'poll_id': pid, 'answer_index': 0})
        assert r.status_code == 200
        assert r.get_json()['success'] is True

    def test_vote_invalid_index(self, client):
        pid = create_poll(client, 'Q', ['A', 'B'])
        r = client.post('/api/vote', json={'poll_id': pid, 'answer_index': 5})
        assert r.status_code == 400

    def test_vote_unknown_poll(self, client):
        r = client.post('/api/vote', json={'poll_id': 9999, 'answer_index': 0})
        assert r.status_code == 404

    def test_vote_recorded_in_stats(self, client):
        pid = create_poll(client, 'Color?', ['Red', 'Blue'])
        sid = create_survey(client, 'S', [pid])

        # Cast 2 votes for Red (index 0), 1 for Blue (index 1)
        client.post('/api/vote', json={'poll_id': pid, 'answer_index': 0})
        client.post('/api/vote', json={'poll_id': pid, 'answer_index': 0})
        client.post('/api/vote', json={'poll_id': pid, 'answer_index': 1})

        r = client.get(f'/api/admin/surveys/{sid}/stats', headers=auth_headers())
        stats = r.get_json()['stats'][0]['stats']
        assert stats['total_votes'] == 3
        assert stats['answer_counts']['0'] == 2
        assert stats['answer_counts']['1'] == 1

    def test_vote_with_session_id(self, client):
        pid = create_poll(client, 'Q?', ['Yes', 'No'])
        r = client.post('/api/vote', json={
            'poll_id': pid, 'answer_index': 0,
            'session_id': 'test-session-uuid-123'
        })
        assert r.status_code == 200
        assert r.get_json()['success'] is True


# ===========================================================================
# EXCEL EXPORT
# ===========================================================================

class TestExcelExport:

    def test_export_returns_xlsx(self, client):
        pid = create_poll(client, 'Export Q?', ['Opt A', 'Opt B'])
        sid = create_survey(client, 'Export Survey', [pid])
        client.post('/api/vote', json={'poll_id': pid, 'answer_index': 0,
                                       'session_id': 'sess-1'})

        r = client.get(f'/api/admin/surveys/{sid}/export', headers=auth_headers())
        assert r.status_code == 200
        ct = r.content_type
        assert 'spreadsheetml' in ct or 'openxmlformats' in ct

    def test_export_per_respondent_rows(self, client):
        """Each session_id must produce exactly one row in the Excel."""
        import openpyxl, io
        pid1 = create_poll(client, 'Q1', ['A', 'B'])
        pid2 = create_poll(client, 'Q2', ['X', 'Y'])
        sid  = create_survey(client, 'Resp Survey', [pid1, pid2])

        # Respondent 1: answers both questions
        client.post('/api/vote', json={'poll_id': pid1, 'answer_index': 0, 'session_id': 'r1'})
        client.post('/api/vote', json={'poll_id': pid2, 'answer_index': 1, 'session_id': 'r1'})
        # Respondent 2: answers only Q1
        client.post('/api/vote', json={'poll_id': pid1, 'answer_index': 1, 'session_id': 'r2'})

        r = client.get(f'/api/admin/surveys/{sid}/export', headers=auth_headers())
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active

        # Row 1 = title, Row 2 = headers, Rows 3+ = respondents
        data_rows = [ws.cell(row=i, column=1).value for i in range(3, ws.max_row + 1)
                     if ws.cell(row=i, column=1).value is not None]
        assert len(data_rows) == 2  # 2 respondents

        # Check answers in row 3 (respondent 1): col3=Q1 answer, col4=Q2 answer
        assert ws.cell(row=3, column=3).value == 'A'   # pid1 index 0
        assert ws.cell(row=3, column=4).value == 'Y'   # pid2 index 1
        # Respondent 2: answered Q1 only, Q2 cell is empty (None or '')
        assert ws.cell(row=4, column=3).value == 'B'           # pid1 index 1
        assert (ws.cell(row=4, column=4).value or '') == ''    # no answer for Q2

    def test_export_unknown_survey_returns_404(self, client):
        r = client.get('/api/admin/surveys/9999/export', headers=auth_headers())
        assert r.status_code == 404
